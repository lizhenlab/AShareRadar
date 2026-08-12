from __future__ import annotations

from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import json
import math
import re
from typing import Final, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models.market_scan import (
    MarketScanResultItem,
    MarketScanResultPage,
    MarketScanResultStatus,
    MarketScanSort,
    MarketScanSortOrder,
)
from app.utils.time import datetime_to_text


XLSX_MEDIA_TYPE: Final = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PUBLISHED_MARKET_SCAN_STATUSES: Final = frozenset({"success", "degraded"})
_DANGEROUS_FORMULA_PREFIXES: Final = frozenset({"=", "+", "-", "@"})
_ILLEGAL_CELL_CHARACTERS: Final = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_MAX_CELL_TEXT_LENGTH: Final = 4_096
_HEADER_FILL: Final = PatternFill(fill_type="solid", fgColor="1F4E78")
_HEADER_FONT: Final = Font(color="FFFFFF", bold=True)
_STATUS_LABELS: Final = {
    "pending": "待处理",
    "success": "有效排名",
    "missing": "数据缺失",
    "skipped": "已跳过",
}
_MODE_LABELS: Final = {
    "official": "盘后正式",
    "intraday": "盘中临时",
    "preopen": "盘前复盘",
}
_SORT_LABELS: Final = {
    "rank": "趋势强度排名",
    "score": "趋势强度",
    "raw_score": "原始得分",
    "trend_score": "趋势分",
    "change_pct": "涨跌幅",
    "amount": "成交额",
    "turnover_rate": "换手率",
    "data_quality_score": "数据质量",
    "alpha_5d": "5日 Alpha",
    "confidence": "置信度",
    "risk": "风险分",
    "tradability": "可交易性",
    "symbol": "股票代码",
}
_RESULT_COLUMNS: Final = (
    ("排名", 9), ("股票代码", 12), ("交易所代码", 15), ("股票名称", 18),
    ("市场", 9), ("上市板块", 18), ("行业", 18), ("上市日期", 13), ("ST", 8), ("新股", 8),
    ("结果状态", 12), ("趋势强度", 14), ("原始得分", 14), ("基础趋势分", 11),
    ("龙头分", 11), ("最新价", 12), ("涨跌幅(%)", 13), ("换手率(%)", 13),
    ("量比", 10), ("成交额(元)", 18), ("数据质量", 12), ("标签", 28),
    ("说明", 36), ("错误", 36), ("行情日期", 13), ("行情时间", 22),
    ("行情来源", 15), ("日K来源", 15), ("复权方式", 12), ("降级原因", 28),
    ("更新时间", 22),
)
_RESULT_LAST_COLUMN: Final = get_column_letter(len(_RESULT_COLUMNS))
_DETAIL_COLUMNS: Final = (
    ("批次 ID", 10), ("股票代码", 12), ("交易所代码", 15), ("最终分", 10),
    ("原始排名分", 15), ("趋势分", 10), ("龙头分", 10), ("数据质量", 12),
    ("龙头基础分", 14), ("趋势增减分", 14), ("规则增减分", 42), ("质量扣分", 12),
    ("扣分前基础分", 16), ("精排综合值", 14), ("精排组成", 48), ("精排扣分", 12),
    ("Tie-break 规则", 34), ("Tie-break 值", 34), ("规则版本", 34), ("规则哈希", 66),
    ("1日Alpha", 12), ("5日Alpha", 12), ("20日Alpha", 12), ("置信度", 12),
    ("风险分", 12), ("可交易性", 12), ("稳健效用", 12), ("均衡效用", 12), ("进取效用", 12),
    ("时点证据状态", 28), ("时点证据摘要", 66),
)
_DETAIL_LAST_COLUMN: Final = get_column_letter(len(_DETAIL_COLUMNS))
_PROBABILITY_HORIZONS: Final = (1, 5, 20)
_PROBABILITY_TARGETS: Final = ("net_excess_positive", "absolute_net_positive")
_PROBABILITY_COLUMNS: Final = (
    ("批次 ID", 10), ("股票代码", 12), ("交易所代码", 15), ("股票名称", 18),
    ("周期(日)", 10), ("目标", 22), ("状态", 20), ("概率", 12),
    ("CI 下限", 12), ("CI 上限", 12), ("CI 水平", 12), ("基础胜率", 12),
    ("版本", 52), ("训练截止", 14), ("Digests", 68), ("局限", 60),
)
_PROBABILITY_LAST_COLUMN: Final = get_column_letter(len(_PROBABILITY_COLUMNS))
_FUTURE_RANGE_COLUMNS: Final = (
    ("研究状态", 18), ("批次 ID", 10), ("股票代码", 15), ("股票名称", 18),
    ("市场", 9), ("行业", 18), ("排名", 9), ("原始得分", 14), ("趋势分", 10),
    ("未来交易日", 12), ("目标交易日", 14), ("固定交易日状态", 18), ("缺失原因", 30),
    ("D日日期", 14), ("D日开盘", 12), ("D日最低", 12),
    ("D日HLC3代理(非VWAP)", 22), ("D日最高", 12), ("D日收盘", 12),
    ("目标日开盘", 12), ("目标日最低", 12), ("目标日HLC3代理(非VWAP)", 24),
    ("目标日最高", 12), ("目标日收盘", 12),
    ("同名最低变化", 14), ("同名HLC3代理变化", 20), ("同名最高变化", 14),
    ("相对D收盘最低", 16), ("相对D收盘HLC3代理", 22),
    ("相对D收盘最高", 16), ("相对D收盘收盘", 16),
    ("D+1开盘入场日", 16), ("D+1开盘参考价", 16),
    ("指定日最低收益", 16), ("指定日HLC3代理收益", 22),
    ("指定日最高收益", 16), ("指定日收盘收益", 16),
    ("累计MAE", 13), ("累计MFE", 13), ("终点收盘收益", 16),
    ("标准化区间宽度", 17), ("区间宽度变化", 16), ("区间重叠率", 14),
    ("更高高点", 12), ("更高低点", 12), ("完整向上缺口", 15), ("完整向下缺口", 15),
    ("执行状态", 18), ("执行不可用原因", 32),
    ("执行入场日", 14), ("执行入场价", 14), ("执行退出日", 14), ("执行退出价", 14),
    ("执行毛收益", 13), ("执行净收益", 13), ("市场基准净收益", 18),
    ("执行净超额收益", 17), ("执行成本拖累", 15),
    ("成本模型版本", 28), ("成本档位", 18), ("执行名义金额", 16),
    ("最大日参与率", 16), ("日线执行模型受限", 20),
    ("概率状态", 18), ("上涨概率上下文", 58), ("时点证据", 58),
    ("目标K线摘要", 66), ("复权方式", 12), ("数据版本", 28), ("K线契约版本", 28),
)
_FUTURE_RANGE_LAST_COLUMN: Final = get_column_letter(len(_FUTURE_RANGE_COLUMNS))


@dataclass(frozen=True)
class MarketScanExportFilters:
    status: MarketScanResultStatus | None = "success"
    market: str | Sequence[str] | None = None
    industry: str | Sequence[str] | None = None
    is_st: bool | None = None
    is_new: bool | None = None
    min_score: int | None = None
    max_score: int | None = None
    min_trend_score: int | None = None
    max_trend_score: int | None = None
    min_change_pct: float | None = None
    max_change_pct: float | None = None
    min_turnover_rate: float | None = None
    max_turnover_rate: float | None = None
    min_amount: float | None = None
    max_amount: float | None = None
    min_data_quality_score: int | None = None
    max_data_quality_score: int | None = None
    min_confidence: float | None = None
    max_risk: float | None = None
    min_tradability: float | None = None
    probability_horizon: Literal[1, 5, 20] = 5
    min_upside_probability: float | None = None
    keyword: str | None = None
    sort: MarketScanSort | Sequence[MarketScanSort] = "rank"
    order: MarketScanSortOrder | Sequence[MarketScanSortOrder] = "asc"

    def normalized(self) -> "MarketScanExportFilters":
        return MarketScanExportFilters(
            status=self.status,
            market=_normalized_filter_values(self.market, maximum=3),
            industry=_normalized_filter_values(self.industry, maximum=20),
            is_st=self.is_st,
            is_new=self.is_new,
            min_score=self.min_score,
            max_score=self.max_score,
            min_trend_score=self.min_trend_score,
            max_trend_score=self.max_trend_score,
            min_change_pct=self.min_change_pct,
            max_change_pct=self.max_change_pct,
            min_turnover_rate=self.min_turnover_rate,
            max_turnover_rate=self.max_turnover_rate,
            min_amount=self.min_amount,
            max_amount=self.max_amount,
            min_data_quality_score=self.min_data_quality_score,
            max_data_quality_score=self.max_data_quality_score,
            min_confidence=self.min_confidence,
            max_risk=self.max_risk,
            min_tradability=self.min_tradability,
            probability_horizon=self.probability_horizon,
            min_upside_probability=self.min_upside_probability,
            keyword=_normalized_filter_text(self.keyword),
            sort=_normalized_sort_values(self.sort, maximum=3),
            order=_normalized_order_values(self.order, maximum=3),
        )


@dataclass(frozen=True)
class MarketScanWorkbookExport:
    content: bytes
    filename: str
    row_count: int


def build_market_scan_workbook(
    page: MarketScanResultPage,
    filters: MarketScanExportFilters,
    *,
    exported_at: datetime,
    future_range_research: dict[str, object] | None = None,
) -> MarketScanWorkbookExport:
    filters = filters.normalized()
    if len(page.items) != page.total:
        raise ValueError("Excel 导出结果读取不完整，请稍后重试")
    workbook = Workbook()
    results_sheet = workbook.active
    if results_sheet is None:
        raise RuntimeError("Excel 工作簿初始化失败")
    results_sheet.title = "榜单"
    _populate_results_sheet(results_sheet, page.items)
    _populate_score_details_sheet(workbook.create_sheet("评分明细"), page.items)
    _populate_probability_sheet(workbook.create_sheet("上涨概率研究"), page)
    _populate_future_range_sheet(
        workbook.create_sheet("未来区间验证"),
        future_range_research,
        run_id=page.run.id,
    )
    _populate_info_sheet(workbook.create_sheet("导出信息"), page, filters, exported_at)
    stream = BytesIO()
    workbook.save(stream)
    return MarketScanWorkbookExport(
        content=stream.getvalue(),
        filename=_export_filename(page, exported_at),
        row_count=page.total,
    )


def _populate_results_sheet(sheet, items: list[MarketScanResultItem]) -> None:
    sheet.append([header for header, _width in _RESULT_COLUMNS])
    for item in items:
        sheet.append(_result_row(item))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_RESULT_LAST_COLUMN}{max(1, len(items) + 1)}"
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, (_header, width) in enumerate(_RESULT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    _format_result_columns(sheet, len(items))
    if items:
        table = Table(displayName="MarketScanResults", ref=f"A1:{_RESULT_LAST_COLUMN}{len(items) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def _result_row(item: MarketScanResultItem) -> list[object]:
    return [
        item.rank, _safe_text(item.code.zfill(6)), _safe_text(item.symbol), _safe_text(item.name),
        _safe_text(item.market), _safe_text(market_scan_board_label(item.code, item.market)),
        _safe_text(item.industry), _safe_text(item.list_date),
        "是" if item.is_st else "否", "是" if item.is_new else "否", _STATUS_LABELS[item.status],
        item.score, item.raw_score, item.trend_score, item.leader_score, item.price,
        item.change_pct, item.turnover_rate, item.volume_ratio, item.amount,
        item.data_quality_score, _safe_text("、".join(item.tags)), _safe_text(item.reason),
        _safe_text(item.error), _safe_text(item.data_date), _safe_text(item.quote_timestamp),
        _safe_text(item.quote_source), _safe_text(item.kline_source), _safe_text(item.adjustment_mode),
        _safe_text("、".join(item.degradation_reasons)), _safe_text(item.updated_at),
    ]


def _format_result_columns(sheet, item_count: int) -> None:
    for row in range(2, item_count + 2):
        sheet.cell(row, 2).number_format = "@"
        for column in (12, 14, 15, 21):
            sheet.cell(row, column).number_format = "0"
        sheet.cell(row, 13).number_format = "0.000000"
        for column in (16, 17, 18, 19):
            sheet.cell(row, column).number_format = "0.00"
        sheet.cell(row, 20).number_format = "#,##0.00"
        for column in (22, 23, 24, 30):
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")


def market_scan_board_label(code: str, market: str) -> str:
    normalized_code = str(code or "").strip()
    normalized_market = str(market or "").strip().upper()
    if normalized_market == "BJ":
        return "北交所"
    if normalized_market == "SH" and normalized_code.startswith(("688", "689")):
        return "科创板"
    if normalized_market == "SZ" and normalized_code.startswith(("300", "301")):
        return "创业板"
    if normalized_market == "SH":
        return "上海A股（主板）"
    if normalized_market == "SZ":
        return "深圳A股（主板）"
    return normalized_market or "板块待确认"


def _populate_score_details_sheet(sheet, items: list[MarketScanResultItem]) -> None:
    sheet.append([header for header, _width in _DETAIL_COLUMNS])
    for item in items:
        sheet.append(_score_detail_row(item))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_DETAIL_LAST_COLUMN}{max(1, len(items) + 1)}"
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, (_header, width) in enumerate(_DETAIL_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    _format_score_detail_columns(sheet, len(items))
    if items:
        table = Table(displayName="MarketScanScoreDetails", ref=f"A1:{_DETAIL_LAST_COLUMN}{len(items) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def _score_detail_row(item: MarketScanResultItem) -> list[object]:
    details = _mapping(item.score_details)
    components = _mapping(details.get("components"))
    leader = _mapping(components.get("leader_score"))
    final_score = _mapping(components.get("final_score"))
    refinement = _mapping(components.get("rank_refinement"))
    ranking = _mapping(details.get("ranking"))
    dimensions = _mapping(components.get("score_dimensions"))
    dimension_scores = _mapping(dimensions.get("scores"))
    utilities = _mapping(dimension_scores.get("decision_utility"))
    evidence = _mapping(dimensions.get("point_in_time_evidence"))
    return [
        item.run_id, _safe_text(item.code.zfill(6)), _safe_text(item.symbol), item.score,
        item.raw_score, item.trend_score, item.leader_score, item.data_quality_score,
        _finite_number(leader.get("base")), _finite_number(leader.get("trend_delta")),
        _safe_json(leader.get("rule_deltas")), _finite_number(final_score.get("quality_penalty")),
        _finite_number(final_score.get("base")), _finite_number(refinement.get("score")),
        _safe_json(refinement.get("weighted_terms")), _finite_number(final_score.get("rank_discount")),
        _safe_json(ranking.get("tie_break")), _safe_json(ranking.get("tie_break_values")),
        _safe_text(details.get("run_rule_version")), _safe_text(details.get("score_spec_hash")),
        _finite_number(dimension_scores.get("alpha_1d")),
        _finite_number(dimension_scores.get("alpha_5d")),
        _finite_number(dimension_scores.get("alpha_20d")),
        _finite_number(dimension_scores.get("confidence")),
        _finite_number(dimension_scores.get("risk")),
        _finite_number(dimension_scores.get("tradability")),
        _finite_number(utilities.get("conservative")),
        _finite_number(utilities.get("balanced")),
        _finite_number(utilities.get("aggressive")),
        _safe_text(evidence.get("status")),
        _safe_text(evidence.get("payload_digest")),
    ]


def _format_score_detail_columns(sheet, item_count: int) -> None:
    for row in range(2, item_count + 2):
        sheet.cell(row, 2).number_format = "@"
        for column in (4, 6, 7, 8):
            sheet.cell(row, column).number_format = "0"
        for column in (5, 9, 10, 12, 13, 14, 16, *range(21, 30)):
            sheet.cell(row, column).number_format = "0.000000"
        for column in (11, 15, 17, 18, 19, 20, 30, 31):
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")


def _populate_probability_sheet(sheet, page: MarketScanResultPage) -> None:
    rows = list(_probability_rows(page))
    sheet.append([header for header, _width in _PROBABILITY_COLUMNS])
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_PROBABILITY_LAST_COLUMN}{max(1, len(rows) + 1)}"
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, (_header, width) in enumerate(_PROBABILITY_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    _format_probability_columns(sheet, len(rows))
    if rows:
        table = Table(displayName="MarketScanProbabilityResearch", ref=f"A1:{_PROBABILITY_LAST_COLUMN}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def _probability_rows(page: MarketScanResultPage) -> Iterator[list[object]]:
    research_horizons = _mapping(_mapping(page.probability_research).get("horizons"))
    for item in page.items:
        item_horizons = _mapping(item.upside_probabilities)
        for horizon in _PROBABILITY_HORIZONS:
            records = _mapping(item_horizons.get(str(horizon)))
            studies = _mapping(research_horizons.get(str(horizon)))
            for target in _PROBABILITY_TARGETS:
                record = _target_mapping(records, target)
                study = _target_mapping(studies, target)
                yield _probability_row(item, horizon, target, record, study)


def _probability_row(
    item: MarketScanResultItem,
    horizon: int,
    target: str,
    record: dict[str, object],
    study: dict[str, object],
) -> list[object]:
    status = _safe_text(record.get("status") or "not_generated")
    probability = _probability_number(record.get("probability")) if status == "calibrated_shadow" else None
    lower, upper, level = _probability_interval(record.get("confidence_interval"), probability)
    base_rate = _probability_number(record.get("base_rate") if "base_rate" in record else study.get("base_rate"))
    versions = _probability_versions(record, study)
    digests = _probability_digests(record, study)
    limitations = record.get("limitations") if "limitations" in record else study.get("limitations")
    return [
        item.run_id, _safe_text(item.code.zfill(6)), _safe_text(item.symbol), _safe_text(item.name),
        horizon, _safe_text(target), status, probability, lower, upper, level, base_rate,
        _safe_json(versions) if versions else "", _safe_text(record.get("training_cutoff") or study.get("training_cutoff")),
        _safe_json(digests) if digests else "", _safe_text("；".join(_string_values(limitations))),
    ]


def _format_probability_columns(sheet, row_count: int) -> None:
    for row in range(2, row_count + 2):
        sheet.cell(row, 2).number_format = "@"
        for column in (8, 9, 10, 11, 12):
            sheet.cell(row, column).number_format = "0.00%"
        for column in (13, 15, 16):
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")


def _populate_future_range_sheet(
    sheet,
    projection: dict[str, object] | None,
    *,
    run_id: int,
) -> None:
    rows = list(_future_range_rows(projection, run_id=run_id))
    sheet.append([header for header, _width in _FUTURE_RANGE_COLUMNS])
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_FUTURE_RANGE_LAST_COLUMN}{max(1, len(rows) + 1)}"
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, (_header, width) in enumerate(_FUTURE_RANGE_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    _format_future_range_columns(sheet, len(rows))
    if rows:
        table = Table(
            displayName="MarketScanFutureRangeResearch",
            ref=f"A1:{_FUTURE_RANGE_LAST_COLUMN}{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def _future_range_rows(
    projection: dict[str, object] | None,
    *,
    run_id: int,
) -> Iterator[list[object]]:
    wrapper = _mapping(projection)
    generation_status = _safe_text(wrapper.get("generation_status") or "not_generated")
    record_page = _mapping(wrapper.get("record_page"))
    research = _mapping(wrapper.get("research"))
    execution_contract = _mapping(
        _mapping(research.get("config")).get("execution_label_contract")
    )
    records = record_page.get("items")
    if not isinstance(records, list) or not records:
        yield _empty_future_range_row(generation_status, run_id)
        return
    for value in records:
        record = _mapping(value)
        offsets = record.get("offsets")
        if not isinstance(offsets, list) or not offsets:
            yield _future_range_row(generation_status, record, {}, execution_contract)
            continue
        for offset in offsets:
            yield _future_range_row(
                generation_status,
                record,
                _mapping(offset),
                execution_contract,
            )


def _empty_future_range_row(generation_status: str, run_id: int) -> list[object]:
    return [generation_status, run_id, *(None for _index in range(len(_FUTURE_RANGE_COLUMNS) - 2))]


def _future_range_row(
    generation_status: str,
    record: dict[str, object],
    offset: dict[str, object],
    execution_contract: dict[str, object],
) -> list[object]:
    d_bar = _mapping(record.get("d_bar"))
    target_bar = _mapping(offset.get("target_bar"))
    level_shift = _mapping(offset.get("level_shift"))
    d_close = _mapping(offset.get("d_close_reference"))
    d1_open = _mapping(offset.get("d1_open_reference"))
    specified = _mapping(d1_open.get("specified_day"))
    cumulative = _mapping(d1_open.get("cumulative_path"))
    interval = _mapping(offset.get("interval_structure"))
    execution = _mapping(offset.get("execution"))
    probability = _mapping(record.get("probability"))
    evidence = _mapping(record.get("source_evidence"))
    return [
        generation_status,
        *_future_range_identity_columns(record, offset),
        *_future_range_bar_columns(d_bar, target_bar),
        *_future_range_outcome_columns(level_shift, d_close, d1_open, specified, cumulative, interval),
        *_future_range_execution_columns(execution, execution_contract),
        *_future_range_context_columns(probability, evidence, offset, d_bar, target_bar),
    ]


def _future_range_identity_columns(
    record: dict[str, object],
    offset: dict[str, object],
) -> list[object]:
    return [
        _finite_number(record.get("run_id")),
        _safe_text(record.get("symbol")),
        _safe_text(record.get("name")),
        _safe_text(record.get("market")),
        _safe_text(record.get("industry")),
        _finite_number(record.get("rank")),
        _finite_number(record.get("raw_score")),
        _finite_number(record.get("trend_score")),
        _finite_number(offset.get("session_offset")),
        _safe_text(offset.get("target_session_date")),
        _safe_text(offset.get("fixed_session_status")),
        _safe_text(offset.get("reason")),
    ]


def _future_range_bar_columns(
    d_bar: dict[str, object],
    target_bar: dict[str, object],
) -> list[object]:
    return [
        _safe_text(d_bar.get("date")),
        _finite_number(d_bar.get("open")),
        _finite_number(d_bar.get("low")),
        _finite_number(d_bar.get("hlc3_proxy")),
        _finite_number(d_bar.get("high")),
        _finite_number(d_bar.get("close")),
        _finite_number(target_bar.get("open")),
        _finite_number(target_bar.get("low")),
        _finite_number(target_bar.get("hlc3_proxy")),
        _finite_number(target_bar.get("high")),
        _finite_number(target_bar.get("close")),
    ]


def _future_range_outcome_columns(
    level_shift: dict[str, object],
    d_close: dict[str, object],
    d1_open: dict[str, object],
    specified: dict[str, object],
    cumulative: dict[str, object],
    interval: dict[str, object],
) -> list[object]:
    return [
        _finite_number(level_shift.get("low")),
        _finite_number(level_shift.get("hlc3_proxy")),
        _finite_number(level_shift.get("high")),
        _finite_number(d_close.get("low")),
        _finite_number(d_close.get("hlc3_proxy")),
        _finite_number(d_close.get("high")),
        _finite_number(d_close.get("close")),
        _safe_text(d1_open.get("entry_date")),
        _finite_number(d1_open.get("entry_price")),
        _finite_number(specified.get("low")),
        _finite_number(specified.get("hlc3_proxy")),
        _finite_number(specified.get("high")),
        _finite_number(specified.get("close")),
        _finite_number(cumulative.get("mae")),
        _finite_number(cumulative.get("mfe")),
        _finite_number(cumulative.get("terminal_close_return")),
        _finite_number(interval.get("normalized_width")),
        _finite_number(interval.get("width_change")),
        _finite_number(interval.get("overlap_ratio")),
        _boolean_label(interval.get("higher_high")),
        _boolean_label(interval.get("higher_low")),
        _boolean_label(interval.get("full_gap_up")),
        _boolean_label(interval.get("full_gap_down")),
    ]


def _future_range_context_columns(
    probability: dict[str, object],
    evidence: dict[str, object],
    offset: dict[str, object],
    d_bar: dict[str, object],
    target_bar: dict[str, object],
) -> list[object]:
    return [
        _safe_text(probability.get("status")),
        _safe_json(probability.get("predictions")),
        _safe_json(evidence),
        _safe_text(offset.get("target_bar_digest")),
        _safe_text(target_bar.get("adjustment_mode") or d_bar.get("adjustment_mode")),
        _safe_text(target_bar.get("data_version") or d_bar.get("data_version")),
        _safe_text(target_bar.get("contract_version") or d_bar.get("contract_version")),
    ]


def _future_range_execution_columns(
    execution: dict[str, object],
    contract: dict[str, object],
) -> list[object]:
    return [
        _safe_text(execution.get("status")),
        _safe_text(execution.get("reason")),
        _safe_text(execution.get("entry_date")),
        _finite_number(execution.get("entry_price")),
        _safe_text(execution.get("exit_date")),
        _finite_number(execution.get("exit_price")),
        _finite_number(execution.get("gross_return")),
        _finite_number(execution.get("net_return")),
        _finite_number(execution.get("market_benchmark_net_return")),
        _finite_number(execution.get("net_excess_return")),
        _finite_number(execution.get("cost_drag")),
        _safe_text(execution.get("cost_model_version") or contract.get("cost_model_version")),
        _safe_text(execution.get("cost_profile_id") or contract.get("cost_profile_id")),
        _finite_number(contract.get("execution_notional")),
        _finite_number(contract.get("max_daily_participation_rate")),
        _boolean_label(execution.get("daily_bar_model_limited")),
    ]


def _format_future_range_columns(sheet, row_count: int) -> None:
    for row in range(2, row_count + 2):
        sheet.cell(row, 3).number_format = "@"
        for column in (7, 9, 10):
            sheet.cell(row, column).number_format = "0"
        sheet.cell(row, 8).number_format = "0.000000"
        for column in (*range(15, 25), 33, 51, 53):
            sheet.cell(row, column).number_format = "0.0000"
        for column in (*range(25, 44), *range(54, 59), 62):
            sheet.cell(row, column).number_format = "0.00%"
        sheet.cell(row, 61).number_format = "#,##0.00"
        for column in (13, 49, 65, 66, 67, 68, 69, 70):
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")


def _boolean_label(value: object) -> str:
    if value is True:
        return "是"
    if value is False:
        return "否"
    return ""


def _target_mapping(values: dict[str, object], target: str) -> dict[str, object]:
    value = values.get(target)
    if value is None and target == "absolute_net_positive":
        value = values.get("net_return_positive")
    return _mapping(value)


def _probability_number(value: object) -> float | None:
    if isinstance(value, bool) or not isinstance(value, int | float):
        return None
    number = float(value)
    return number if math.isfinite(number) and 0 <= number <= 1 else None


def _probability_interval(value: object, probability: float | None) -> tuple[float | None, float | None, float | None]:
    if probability is None:
        return None, None, None
    if isinstance(value, dict):
        lower, upper, level = value.get("lower"), value.get("upper"), value.get("level", 0.95)
    elif isinstance(value, list | tuple) and len(value) == 2:
        lower, upper, level = value[0], value[1], 0.95
    else:
        return None, None, None
    lower_number = _probability_number(lower)
    upper_number = _probability_number(upper)
    level_number = _probability_number(level)
    if lower_number is None or upper_number is None or level_number is None:
        return None, None, None
    if not lower_number <= probability <= upper_number:
        return None, None, None
    return lower_number, upper_number, level_number


def _probability_versions(record: dict[str, object], study: dict[str, object]) -> dict[str, object]:
    versions = {**_mapping(study.get("versions")), **_mapping(record.get("versions"))}
    flat_keys = {
        "model": "model_version", "calibrator": "calibrator_version", "feature": "feature_version",
        "label": "label_version", "cost_model": "cost_model_version", "benchmark": "benchmark_version",
    }
    for name, key in flat_keys.items():
        value = record.get(key) or study.get(key)
        if value is not None:
            versions[name] = value
    return {key: value for key, value in versions.items() if value not in (None, "")}


def _probability_digests(record: dict[str, object], study: dict[str, object]) -> dict[str, object]:
    digests = {**_mapping(study.get("digests")), **_mapping(record.get("digests"))}
    flat_keys = {
        "input": ("input_digest",), "model": ("model_digest",),
        "calibrator": ("calibrator_digest",), "baseline": ("baseline_digest",),
        "evidence": ("artifact_id", "evidence_digest"),
    }
    for name, keys in flat_keys.items():
        value = next((source.get(key) for source in (record, study) for key in keys if source.get(key)), None)
        if value is not None:
            digests[name] = value
    return {key: value for key, value in digests.items() if value not in (None, "")}


def _string_values(value: object) -> list[str]:
    if isinstance(value, str):
        return [value]
    if not isinstance(value, list | tuple):
        return []
    return [item for item in value if isinstance(item, str) and item]


def _populate_info_sheet(sheet, page: MarketScanResultPage, filters: MarketScanExportFilters, exported_at: datetime) -> None:
    run = page.run
    rows = (
        ("项目", "AShareRadar"), ("导出类型", "全市场A股榜单"), ("批次 ID", run.id),
        ("榜单类型", _MODE_LABELS[run.mode]), ("批次状态", run.status), ("触发方式", run.trigger),
        ("批次基准时间", run.as_of), ("行情日期", run.quote_date),
        ("日K截止日", run.data_date), ("扫描完成时间", _text_or(run.finished_at, "--")), ("规则版本", run.rule_version),
        ("股票池范围", run.scope), ("股票池来源", _text_or(run.stock_pool_source, "--")),
        ("有效覆盖率", f"{run.coverage_pct:.2f}%"),
        ("筛选状态", _status_filter_label(filters.status)),
        ("筛选市场", _filter_values_label(filters.market, "全部市场")),
        ("筛选行业", _filter_values_label(filters.industry, "不限")),
        ("ST 条件", _boolean_filter_label(filters.is_st, "仅 ST", "排除 ST")),
        ("新股条件", _boolean_filter_label(filters.is_new, "仅新股", "排除新股")),
        ("趋势强度范围", _range_filter_label(filters.min_score, filters.max_score)),
        ("趋势分范围", _range_filter_label(filters.min_trend_score, filters.max_trend_score)),
        ("涨跌幅范围", _range_filter_label(filters.min_change_pct, filters.max_change_pct)),
        ("换手率范围", _range_filter_label(filters.min_turnover_rate, filters.max_turnover_rate)),
        ("成交额范围", _range_filter_label(filters.min_amount, filters.max_amount)),
        ("数据质量范围", _range_filter_label(filters.min_data_quality_score, filters.max_data_quality_score)),
        ("最低置信度", _range_filter_label(filters.min_confidence, None)),
        ("最高风险分", _range_filter_label(None, filters.max_risk)),
        ("最低可交易性", _range_filter_label(filters.min_tradability, None)),
        ("上涨概率周期", f"{filters.probability_horizon}日"),
        ("最低上涨概率", _probability_filter_label(filters.min_upside_probability)),
        ("搜索关键词", _text_or(filters.keyword, "无")),
        ("排序", _sort_filter_label(filters.sort, filters.order)),
        ("导出条数", page.total), ("导出时间", datetime_to_text(exported_at)),
        ("数据说明", "仅导出已持久化榜单快照，不会重新获取行情或重新计算。"),
    )
    sheet.append(["字段", "内容"])
    for key, value in rows:
        sheet.append([key, _safe_text(value) if isinstance(value, str) else value])
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 72
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet["A"]:
        cell.font = Font(bold=True)


def _safe_text(value: object | None) -> str:
    text = _ILLEGAL_CELL_CHARACTERS.sub("", str("" if value is None else value))
    if len(text) > _MAX_CELL_TEXT_LENGTH:
        text = f"{text[:_MAX_CELL_TEXT_LENGTH - 3]}..."
    if text.lstrip()[:1] in _DANGEROUS_FORMULA_PREFIXES:
        return f"'{text}"
    return text


def _safe_json(value: object) -> str:
    if value is None:
        return ""
    try:
        rendered = json.dumps(value, ensure_ascii=False, allow_nan=False, sort_keys=True, separators=(",", ":"))
    except (TypeError, ValueError):
        return ""
    return _safe_text(rendered)


def _mapping(value: object) -> dict[str, object]:
    return value if isinstance(value, dict) else {}


def _finite_number(value: object) -> int | float | None:
    if isinstance(value, bool) or not isinstance(value, int | float):
        return None
    return value


def _normalized_filter_text(value: str | None) -> str | None:
    normalized = " ".join((value or "").split()).strip()
    return normalized or None


def _normalized_filter_values(value: str | Sequence[str] | None, *, maximum: int) -> tuple[str, ...]:
    candidates = [value] if isinstance(value, str) else list(value or ())
    normalized = tuple(dict.fromkeys(" ".join(str(item).split()).strip() for item in candidates))
    return tuple(item for item in normalized if item)[:maximum]


def _normalized_sort_values(
    value: MarketScanSort | Sequence[MarketScanSort],
    *,
    maximum: int,
) -> tuple[MarketScanSort, ...]:
    return (value,) if isinstance(value, str) else tuple(value)[:maximum]


def _normalized_order_values(
    value: MarketScanSortOrder | Sequence[MarketScanSortOrder],
    *,
    maximum: int,
) -> tuple[MarketScanSortOrder, ...]:
    return (value,) if isinstance(value, str) else tuple(value)[:maximum]


def _boolean_filter_label(value: bool | None, true_label: str, false_label: str) -> str:
    if value is None:
        return "不限"
    return true_label if value else false_label


def _status_filter_label(value: MarketScanResultStatus | None) -> str:
    return "全部状态" if value is None else _STATUS_LABELS[value]


def _quality_filter_label(value: int | None) -> int | str:
    return value if value is not None else "不限"


def _filter_values_label(value: str | Sequence[str] | None, fallback: str) -> str:
    normalized = _normalized_filter_values(value, maximum=20)
    return "、".join(normalized) if normalized else fallback


def _range_filter_label(minimum: int | float | None, maximum: int | float | None) -> str:
    if minimum is None and maximum is None:
        return "不限"
    if maximum is None:
        return f"≥ {minimum}"
    if minimum is None:
        return f"≤ {maximum}"
    return f"{minimum if minimum is not None else '--'} ～ {maximum if maximum is not None else '--'}"


def _probability_filter_label(value: float | None) -> str:
    return "不限" if value is None else f"≥ {value:.1%}（仅已校准 Shadow）"


def _sort_filter_label(
    sort: MarketScanSort | Sequence[MarketScanSort],
    order: MarketScanSortOrder | Sequence[MarketScanSortOrder],
) -> str:
    sorts = (sort,) if isinstance(sort, str) else tuple(sort)
    orders = (order,) if isinstance(order, str) else tuple(order)
    return " → ".join(
        f"{_SORT_LABELS[field]}（{_sort_order_label(direction)}）"
        for field, direction in zip(sorts, orders, strict=False)
    )


def _sort_order_label(value: MarketScanSortOrder) -> str:
    return "升序" if value == "asc" else "降序"


def _text_or(value: str | None, fallback: str) -> str:
    return value if value else fallback


def _export_filename(page: MarketScanResultPage, exported_at: datetime) -> str:
    date = page.run.quote_date or page.run.data_date or exported_at.date().isoformat()
    return f"AShareRadar-market-scan-{date}-{page.run.mode}-run-{page.run.id}.xlsx"


__all__ = [
    "PUBLISHED_MARKET_SCAN_STATUSES",
    "XLSX_MEDIA_TYPE",
    "MarketScanExportFilters",
    "MarketScanWorkbookExport",
    "build_market_scan_workbook",
    "market_scan_board_label",
]
