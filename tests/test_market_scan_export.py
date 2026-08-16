from __future__ import annotations

from copy import deepcopy
from datetime import datetime
from io import BytesIO
import json

from openpyxl import load_workbook
import pytest

from app.models.market_scan import (
    MARKET_SCAN_TOP100_REFRESH_SCOPE,
    MarketScanResultItem,
    MarketScanResultPage,
    MarketScanRun,
)
from app.services.market_scan_export import (
    MarketScanExportFilters,
    build_market_scan_workbook,
    market_scan_board_label,
)
from app.services.market_scan_manager import MarketScanManager
from app.services.market_scan_probability_store import ProbabilityResearchUnavailable
from app.services.market_scan_universe import FULL_MARKET_SCOPE


EXPORTED_AT = datetime.fromisoformat("2026-07-29T16:20:30+08:00")


def test_workbook_contains_complete_ranked_snapshot_and_audit_metadata() -> None:
    page = _page([_item()])
    filters = MarketScanExportFilters(
        market=("SZ", "SH"),
        industry=("银行", "半导体"),
        is_st=False,
        is_new=False,
        min_score=60,
        max_score=95,
        min_trend_score=50,
        max_trend_score=90,
        min_change_pct=-2.5,
        max_change_pct=9.5,
        min_turnover_rate=1,
        max_turnover_rate=30,
        min_amount=1_000_000,
        max_amount=500_000_000,
        min_data_quality_score=80,
        max_data_quality_score=99,
        min_confidence=70,
        max_risk=40,
        min_tradability=60,
        keyword="000001",
        sort=("amount", "score", "symbol"),
        order=("desc", "desc", "asc"),
    )

    exported = build_market_scan_workbook(page, filters, exported_at=EXPORTED_AT)
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    results = workbook["榜单"]
    info = {row[0].value: row[1].value for row in workbook["导出信息"].iter_rows(min_row=2)}

    assert workbook.sheetnames == ["榜单", "评分明细", "上涨概率研究", "未来区间验证", "导出信息"]
    assert exported.filename == "AShareRadar-market-scan-2026-07-29-official-run-12.xlsx"
    assert exported.row_count == 1
    assert results.freeze_panes == "A2"
    assert results.auto_filter.ref == "A1:AE2"
    assert results["A1"].value == "排名"
    assert results["B2"].value == "000001"
    assert results["B2"].number_format == "@"
    assert results["D2"].value == "'=HYPERLINK(\"https://example.invalid\")"
    assert results["F2"].value == "深圳A股（主板）"
    assert results["G2"].value == "'+银行"
    assert results["W2"].value == "'@趋势和量价配合"
    assert results["X2"].value is None
    assert results["AD2"].value == "'-provider_reason"
    assert all(cell.data_type != "f" for row in results.iter_rows() for cell in row)
    details = workbook["评分明细"]
    assert details["B2"].value == "000001"
    assert details["B2"].number_format == "@"
    assert details["D2"].value == 92
    assert details["E2"].value == pytest.approx(91.123456)
    assert details["K2"].value == '{"amount":2.5}'
    assert details["P2"].value == pytest.approx(-0.276544)
    assert details["Q2"].value == '[["raw_score","desc"],["symbol","asc"]]'
    assert all(cell.data_type != "f" for row in details.iter_rows() for cell in row)
    assert info["批次 ID"] == 12
    assert info["榜单类型"] == "盘后正式"
    assert info["筛选市场"] == "SZ、SH"
    assert info["筛选行业"] == "银行、半导体"
    assert info["ST 条件"] == "排除 ST"
    assert info["趋势强度范围"] == "60 ～ 95"
    assert info["最低置信度"] == "≥ 70"
    assert info["最高风险分"] == "≤ 40"
    assert info["最低可交易性"] == "≥ 60"
    assert info["趋势分范围"] == "50 ～ 90"
    assert info["涨跌幅范围"] == "'-2.5 ～ 9.5"
    assert info["换手率范围"] == "1 ～ 30"
    assert info["成交额范围"] == "1000000 ～ 500000000"
    assert info["数据质量范围"] == "80 ～ 99"
    assert info["排序"] == "成交额（降序） → 趋势强度（降序） → 股票代码（升序）"
    assert info["筛选合同"] == "screen-spec-v2"
    assert len(info["筛选摘要"]) == 64
    assert info["快照摘要"] == "a" * 64
    exported_spec = json.loads(info["筛选合同 JSON"])
    assert exported_spec["ranges"]["confidence"] == {"max": None, "min": 70.0}
    assert exported_spec["keyword"] == "000001"
    assert exported_spec["sort"] == [
        {"field": "amount", "order": "desc"},
        {"field": "score", "order": "desc"},
        {"field": "symbol", "order": "asc"},
    ]
    assert info["导出条数"] == 1
    assert info["数据说明"] == "仅导出已持久化榜单快照，不会重新获取行情或重新计算。"


def test_score_detail_export_reads_v5_continuous_trend_without_losing_v4_columns() -> None:
    item = _item()
    details = deepcopy(item.score_details)
    components = details["components"]
    components["continuous_trend"] = components.pop("rank_refinement")
    components["continuous_trend"]["score"] = 0.662917
    final_score = components["final_score"]
    final_score.pop("rank_discount")
    final_score["continuous_trend_adjustment"] = 1.303336
    workbook = load_workbook(
        BytesIO(
            build_market_scan_workbook(
                _page([item.model_copy(update={"score_details": details})]),
                MarketScanExportFilters(),
                exported_at=EXPORTED_AT,
            ).content
        )
    )
    sheet = workbook["评分明细"]

    assert sheet["N1"].value == "连续趋势/旧精排值"
    assert sheet["N2"].value == pytest.approx(0.662917)
    assert sheet["P2"].value == pytest.approx(1.303336)


def test_workbook_supports_an_empty_filtered_result_without_an_invalid_table() -> None:
    page = _page([])

    exported = build_market_scan_workbook(page, MarketScanExportFilters(), exported_at=EXPORTED_AT)
    workbook = load_workbook(BytesIO(exported.content))

    assert exported.row_count == 0
    assert workbook["榜单"].max_row == 1
    assert workbook["榜单"].tables == {}
    assert workbook["榜单"].auto_filter.ref == "A1:AE1"
    assert workbook["上涨概率研究"].tables == {}
    assert workbook["上涨概率研究"].auto_filter.ref == "A1:T1"


def test_probability_research_sheet_exports_only_available_probabilities_with_explicit_semantics() -> None:
    probabilities = {
        "5": {
            "net_excess_positive": {
                "status": "calibrated_shadow", "probability": 0.612,
                "calibration_bias_interval": {
                    "lower": -0.052, "upper": 0.048, "level": 0.95,
                    "method": "date_block_bootstrap_signed_calibration_bias",
                    "semantics": "signed_observed_rate_minus_probability_bias",
                },
                "calibration_adjusted_probability_interval": {
                    "lower": 0.56, "upper": 0.66, "level": 0.95,
                    "method": "date_block_bootstrap_calibration_offset",
                    "semantics": "calibration_adjusted_probability_interval_not_individual_outcome_interval",
                },
                "base_rate": 0.514, "model_version": "record-model-v1",
                "input_digest": "record-input", "training_cutoff": "2026-07-15",
                "limitations": ["=shadow_only"],
            },
            "absolute_net_positive": {
                "status": "insufficient_data", "probability": 0.5,
                "confidence_interval": [0.4, 0.6], "base_rate": 0.48,
                "limitations": ["insufficient_independent_dates"],
            },
        },
    }
    research: dict[str, object] = {
        "run_binding": {"binding_status": "verified", "legacy": False},
        "horizons": {
            "1": {"net_excess_positive": {"status": "calibrated_shadow", "base_rate": 0.51}},
            "5": {
                "net_excess_positive": {
                    "status": "calibrated_shadow", "base_rate": 0.514,
                    "versions": {"model": "study-model-v1", "feature": "feature-v1", "label": "label-v1", "cost_model": "cost-v1"},
                    "digests": {"input": "study-input", "model": "study-model-digest", "calibrator": "study-calibrator-digest"},
                    "training_cutoff": "2026-07-14", "limitations": ["study_shadow_only"],
                },
                "absolute_net_positive": {"status": "insufficient_data", "base_rate": 0.48},
            },
        },
    }
    item = _item().model_copy(update={"upside_probabilities": probabilities})
    page = _page([item], probability_research=research)

    workbook = load_workbook(BytesIO(build_market_scan_workbook(page, MarketScanExportFilters(), exported_at=EXPORTED_AT).content))
    sheet = workbook["上涨概率研究"]
    rows = {(row[4], row[6]): row for row in sheet.iter_rows(min_row=2, values_only=True)}

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:T2"
    assert "MarketScanProbabilityResearch" in sheet.tables
    assert len(rows) == 1
    calibrated = rows[(5, "net_excess_positive")]
    assert calibrated[5] == "D+6"
    assert calibrated[7] == "calibrated_shadow"
    assert calibrated[8:14] == pytest.approx((0.612, -0.052, 0.048, 0.56, 0.66, 0.95))
    assert calibrated[14] == "有符号偏差与群体校准调整概率区间；不是个股结果区间"
    assert calibrated[15] == pytest.approx(0.514)
    assert json.loads(calibrated[16]) == {
        "cost_model": "cost-v1", "feature": "feature-v1", "label": "label-v1", "model": "record-model-v1",
    }
    assert calibrated[17] == "2026-07-15"
    assert json.loads(calibrated[18]) == {
        "calibrator": "study-calibrator-digest", "input": "record-input", "model": "study-model-digest",
    }
    assert calibrated[19] == "'=shadow_only"
    assert all(sheet.cell(2, column).number_format == "0.00%" for column in (9, 10, 11, 12, 13, 14, 16))
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


def test_probability_research_sheet_keeps_legacy_probabilities_blank() -> None:
    page = _page([_item()])
    workbook = load_workbook(BytesIO(build_market_scan_workbook(page, MarketScanExportFilters(), exported_at=EXPORTED_AT).content))

    rows = list(workbook["上涨概率研究"].iter_rows(min_row=2, values_only=True))
    assert rows == []


@pytest.mark.parametrize(
    ("bias", "adjusted", "expected"),
    (
        (
            {"lower": 0.10, "upper": 0.20, "level": 0.95, "method": "date_block_bootstrap_signed_calibration_bias", "semantics": "signed_observed_rate_minus_probability_bias"},
            {"lower": 0.70, "upper": 0.80, "level": 0.95, "method": "date_block_bootstrap_calibration_offset", "semantics": "calibration_adjusted_probability_interval_not_individual_outcome_interval"},
            (0.10, 0.20, 0.70, 0.80, 0.95),
        ),
        (None, None, (None, None, None, None, None)),
        (
            {"lower": "invalid", "upper": 0.1, "level": 0.95, "method": "date_block_bootstrap_signed_calibration_bias", "semantics": "signed_observed_rate_minus_probability_bias"},
            {"lower": 0.5, "upper": 0.7, "level": 0.95, "method": "date_block_bootstrap_calibration_offset", "semantics": "calibration_adjusted_probability_interval_not_individual_outcome_interval"},
            (None, None, None, None, None),
        ),
    ),
)
def test_probability_export_only_emits_valid_calibration_bias_intervals(
    bias: object,
    adjusted: object,
    expected: tuple[float | None, float | None, float | None, float | None, float | None],
) -> None:
    record = {
        "status": "calibrated_shadow",
        "probability": 0.60,
        "calibration_bias_interval": bias,
        "calibration_adjusted_probability_interval": adjusted,
    }
    probabilities = {"5": {"net_excess_positive": record}}
    research = {
        "run_binding": {"binding_status": "verified", "legacy": False},
        "horizons": {"5": {"net_excess_positive": {"status": "calibrated_shadow"}}},
    }
    item = _item().model_copy(update={"upside_probabilities": probabilities})
    page = _page([item], probability_research=research)

    workbook = load_workbook(
        BytesIO(
            build_market_scan_workbook(
                page,
                MarketScanExportFilters(),
                exported_at=EXPORTED_AT,
            ).content
        )
    )
    rows = list(workbook["上涨概率研究"].iter_rows(min_row=2, values_only=True))

    assert len(rows) == 1
    assert rows[0][8] == pytest.approx(0.60)
    if expected[0] is None:
        assert rows[0][9:14] == expected
    else:
        assert rows[0][9:14] == pytest.approx(expected)


def test_probability_export_rejects_boolean_probability_as_missing_evidence() -> None:
    probabilities = {
        "5": {
            "net_excess_positive": {
                "status": "calibrated_shadow",
                "probability": True,
                "calibration_bias_interval": [0.5, 0.7],
            }
        }
    }
    research = {
        "run_binding": {"binding_status": "verified", "legacy": False},
        "horizons": {"5": {"net_excess_positive": {"status": "calibrated_shadow"}}},
    }
    item = _item().model_copy(update={"upside_probabilities": probabilities})

    workbook = load_workbook(
        BytesIO(
            build_market_scan_workbook(
                _page([item], probability_research=research),
                MarketScanExportFilters(),
                exported_at=EXPORTED_AT,
            ).content
        )
    )

    assert list(workbook["上涨概率研究"].iter_rows(min_row=2, values_only=True)) == []


def test_future_range_sheet_exports_fixed_sessions_and_marks_hlc3_as_proxy() -> None:
    projection = _future_range_projection()
    workbook = load_workbook(
        BytesIO(
            build_market_scan_workbook(
                _page([_item()]),
                MarketScanExportFilters(),
                exported_at=EXPORTED_AT,
                future_range_research=projection,
            ).content
        )
    )
    sheet = workbook["未来区间验证"]
    headers = [cell.value for cell in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert "D日HLC3代理(非VWAP)" in headers
    assert "目标日HLC3代理(非VWAP)" in headers
    assert "指定日HLC3代理收益" in headers
    assert sheet.freeze_panes == "A2"
    assert "执行净收益" in headers
    assert "执行净超额收益" in headers
    assert "成本模型版本" in headers
    assert len(rows) == 3
    available, modelled, unavailable = rows
    assert available[0:13] == (
        "ready", 12, "000001.SZ", "平安银行", "SZ", "银行", 1, 91.123456,
        88, 1, "2026-07-30", "available", None,
    )
    assert available[14:24] == pytest.approx(
        (10.0, 9.8, 10.0, 10.2, 10.1, 10.2, 10.0, 10.2, 10.4, 10.3)
    )
    assert available[24:31] == pytest.approx((0.02, 0.02, 0.0196, -0.0099, 0.0099, 0.0297, 0.0198))
    assert available[31] == "2026-07-30"
    assert available[32] == pytest.approx(10.2)
    assert available[33:43] == pytest.approx((-0.0196, 0.0, 0.0196, 0.0098, -0.0196, 0.0196, 0.0098, 0.0385, 0.0, 0.5))
    assert available[43:47] == ("是", "是", "否", "否")
    assert available[47:53] == (
        "data_unavailable", "A_share_T_plus_1_no_same_session_exit",
        "2026-07-30", None, None, None,
    )
    assert all(value is None for value in available[53:58])
    assert available[63] == "calibrated_shadow"
    assert json.loads(available[64]) == [{"horizon": 1, "probability": 0.61, "target": "net_excess_positive"}]
    assert available[67:70] == ("qfq", "daily-v1", "kline-v1")
    assert modelled[47:53] == (
        "modelled", None, "2026-07-30", 10.2, "2026-07-31", 10.5,
    )
    assert modelled[53:58] == pytest.approx((0.0294, 0.0278, 0.01, 0.0178, 0.0016))
    assert modelled[58:63] == ("ashare-cost-v1", "base", 100000, 0.01, "是")
    assert unavailable[11] == "unavailable"
    assert unavailable[12] == "suspended_or_missing_bar"
    assert all(value is None for value in unavailable[19:43])
    assert unavailable[47] == "data_unavailable"
    assert unavailable[48] == "fixed_path_bar_missing_no_forward_shift"
    assert unavailable[50] is None
    assert unavailable[51] == "2026-08-01"
    assert all(value is None for value in unavailable[52:58])
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


@pytest.mark.parametrize(
    ("code", "market", "expected"),
    (
        ("600519", "SH", "上海A股（主板）"),
        ("688981", "SH", "科创板"),
        ("000001", "SZ", "深圳A股（主板）"),
        ("300750", "SZ", "创业板"),
        ("920066", "BJ", "北交所"),
    ),
)
def test_export_board_labels_are_explicit(code: str, market: str, expected: str) -> None:
    assert market_scan_board_label(code, market) == expected


def test_workbook_rejects_a_truncated_result_page() -> None:
    page = _page([_item()]).model_copy(update={"total": 2})

    with pytest.raises(ValueError, match="读取不完整"):
        build_market_scan_workbook(page, MarketScanExportFilters(), exported_at=EXPORTED_AT)


def test_manager_exports_only_published_runs_and_forwards_every_filter() -> None:
    page = _page([_item()])
    filters = MarketScanExportFilters(
        status=None,
        market=("SZ", "SH"),
        industry=("  银行   服务 ", "电力"),
        is_st=False,
        min_score=60,
        max_score=98,
        min_trend_score=50,
        max_trend_score=95,
        min_change_pct=-3,
        max_change_pct=10,
        min_turnover_rate=1,
        max_turnover_rate=25,
        min_amount=1_000_000,
        max_amount=900_000_000,
        min_data_quality_score=70,
        max_data_quality_score=100,
        min_confidence=75,
        max_risk=35,
        min_tradability=65,
        keyword=" 000001   平安 ",
        sort=("score", "amount", "symbol"),
        order=("desc", "desc", "asc"),
    )
    manager = object.__new__(MarketScanManager)
    calls: list[tuple[int, dict[str, object]]] = []
    manager.run = lambda run_id: page.run  # type: ignore[method-assign]
    manager.results = lambda run_id, **kwargs: calls.append((run_id, kwargs)) or page  # type: ignore[method-assign]
    manager._now = lambda: EXPORTED_AT

    exported = manager.export_results(page.run.id, filters=filters)

    assert exported.row_count == 1
    assert calls == [
        (
            page.run.id,
            {
                "page": 1,
                "page_size": 1,
                "status": None,
                "market": ("SZ", "SH"),
                "industry": ("银行 服务", "电力"),
                "is_st": False,
                "is_new": None,
                "min_score": 60,
                "max_score": 98,
                "min_trend_score": 50,
                "max_trend_score": 95,
                "min_change_pct": -3,
                "max_change_pct": 10,
                "min_turnover_rate": 1,
                "max_turnover_rate": 25,
                "min_amount": 1_000_000,
                "max_amount": 900_000_000,
                "min_data_quality_score": 70,
                    "max_data_quality_score": 100,
                    "min_confidence": 75,
                    "max_risk": 35,
                        "min_tradability": 65,
                    "probability_horizon": 5,
                    "min_upside_probability": None,
                    "keyword": "000001 平安",
                "sort": ("score", "amount", "symbol"),
                "order": ("desc", "desc", "asc"),
            },
        )
    ]

    manager.run = lambda run_id: page.run.model_copy(update={"status": "running"})  # type: ignore[method-assign]
    with pytest.raises(ValueError, match="只有已发布"):
        manager.export_results(page.run.id, filters=filters)


def test_top100_refresh_export_is_rejected_before_research_artifacts_are_read() -> None:
    page = _page([_item()])
    manager = object.__new__(MarketScanManager)
    future_range_calls: list[int] = []

    class _FutureRangeStore:
        def export_projection(self, run_id: int) -> dict[str, object]:
            future_range_calls.append(run_id)
            return _future_range_projection()

    top100 = page.run.model_copy(update={"scope": MARKET_SCAN_TOP100_REFRESH_SCOPE})
    manager.run = lambda _run_id: top100  # type: ignore[method-assign]
    manager.results = lambda _run_id, **_kwargs: page  # type: ignore[method-assign]
    manager._future_range_store = _FutureRangeStore()  # type: ignore[assignment]
    manager._now = lambda: EXPORTED_AT

    with pytest.raises(ProbabilityResearchUnavailable, match="盘后正式全市场"):
        manager.export_results(top100.id, filters=MarketScanExportFilters())
    assert future_range_calls == []


def _page(
    items: list[MarketScanResultItem],
    *,
    probability_research: dict[str, object] | None = None,
) -> MarketScanResultPage:
    return MarketScanResultPage(
        run=_run(),
        items=items,
        total=len(items),
        page=1,
        page_size=max(1, len(items)),
        page_count=1 if items else 0,
        probability_research=probability_research,
    )


def _future_range_projection() -> dict[str, object]:
    d_bar = {
        "date": "2026-07-29", "open": 10.0, "low": 9.8, "hlc3_proxy": 10.0,
        "high": 10.2, "close": 10.1, "adjustment_mode": "qfq",
        "data_version": "daily-v1", "contract_version": "kline-v1",
    }
    return {
        "schema_version": "market-scan-future-range-api-v1",
        "generation_status": "ready",
        "artifact": {
            "schema_version": "market-scan-future-range-artifact-v1",
            "generated_at": "2026-08-01T16:00:00+08:00",
            "integrity_digest": "a" * 64,
        },
        "research": {
            "status": "ok", "record_count": 1,
            "config": {
                "execution_label_contract": {
                    "cost_model_version": "ashare-cost-v1", "cost_profile_id": "base",
                    "execution_notional": 100_000, "max_daily_participation_rate": 0.01,
                }
            },
        },
        "record_page": {
            "page": 1, "page_size": 1, "total": 1, "page_count": 1,
            "session_offset": None, "symbol": None,
            "items": [{
                "run_id": 12, "symbol": "000001.SZ", "name": "平安银行", "market": "SZ",
                "industry": "银行", "rank": 1, "raw_score": 91.123456, "trend_score": 88,
                "d_bar": d_bar,
                "source_evidence": {"status": "verified", "payload_digest": "source-digest"},
                "probability": {
                    "status": "calibrated_shadow",
                    "predictions": [{"target": "net_excess_positive", "horizon": 1, "probability": 0.61}],
                },
                "offsets": [
                    {
                        "session_offset": 1, "target_session_date": "2026-07-30",
                        "fixed_session_status": "available", "reason": None,
                        "target_bar": {
                            "date": "2026-07-30", "open": 10.2, "low": 10.0,
                            "hlc3_proxy": 10.2, "high": 10.4, "close": 10.3,
                            "adjustment_mode": "qfq", "data_version": "daily-v1",
                            "contract_version": "kline-v1",
                        },
                        "target_bar_digest": "target-digest",
                        "level_shift": {"low": 0.02, "hlc3_proxy": 0.02, "high": 0.0196},
                        "d_close_reference": {"low": -0.0099, "hlc3_proxy": 0.0099, "high": 0.0297, "close": 0.0198},
                        "d1_open_reference": {
                            "entry_date": "2026-07-30", "entry_price": 10.2,
                            "specified_day": {"low": -0.0196, "hlc3_proxy": 0.0, "high": 0.0196, "close": 0.0098},
                            "cumulative_path": {"mae": -0.0196, "mfe": 0.0196, "terminal_close_return": 0.0098},
                        },
                        "interval_structure": {
                            "normalized_width": 0.0385, "width_change": 0.0,
                            "overlap_ratio": 0.5, "higher_high": True, "higher_low": True,
                            "full_gap_up": False, "full_gap_down": False,
                        },
                        "execution": {
                            "status": "data_unavailable",
                            "reason": "A_share_T_plus_1_no_same_session_exit",
                            "entry_date": "2026-07-30", "entry_price": None,
                            "exit_date": None, "exit_price": None,
                            "gross_return": None, "net_return": None,
                            "market_benchmark_net_return": None, "net_excess_return": None,
                            "cost_drag": None, "daily_bar_model_limited": None,
                            "cost_model_version": "ashare-cost-v1", "cost_profile_id": "base",
                        },
                    },
                    {
                        "session_offset": 2, "target_session_date": "2026-07-31",
                        "fixed_session_status": "available", "reason": None,
                        "target_bar": {
                            "date": "2026-07-31", "open": 10.3, "low": 10.1,
                            "hlc3_proxy": 10.4, "high": 10.6, "close": 10.5,
                            "adjustment_mode": "qfq", "data_version": "daily-v1",
                            "contract_version": "kline-v1",
                        },
                        "target_bar_digest": "target-digest-2",
                        "level_shift": {"low": 0.0306, "hlc3_proxy": 0.04, "high": 0.0392},
                        "d_close_reference": {"low": 0.0, "hlc3_proxy": 0.0297, "high": 0.0495, "close": 0.0396},
                        "d1_open_reference": {
                            "entry_date": "2026-07-30", "entry_price": 10.2,
                            "specified_day": {"low": -0.0098, "hlc3_proxy": 0.0196, "high": 0.0392, "close": 0.0294},
                            "cumulative_path": {"mae": -0.0196, "mfe": 0.0392, "terminal_close_return": 0.0294},
                        },
                        "interval_structure": {
                            "normalized_width": 0.0481, "width_change": 0.0096,
                            "overlap_ratio": 0.1667, "higher_high": True, "higher_low": True,
                            "full_gap_up": False, "full_gap_down": False,
                        },
                        "execution": {
                            "status": "modelled", "reason": None,
                            "entry_date": "2026-07-30", "entry_price": 10.2,
                            "exit_date": "2026-07-31", "exit_price": 10.5,
                            "gross_return": 0.0294, "net_return": 0.0278,
                            "market_benchmark_net_return": 0.01, "net_excess_return": 0.0178,
                            "cost_drag": 0.0016, "daily_bar_model_limited": True,
                            "cost_model_version": "ashare-cost-v1", "cost_profile_id": "base",
                        },
                    },
                    {
                        "session_offset": 3, "target_session_date": "2026-08-01",
                        "fixed_session_status": "unavailable", "reason": "suspended_or_missing_bar",
                        "target_bar": None, "target_bar_digest": None, "level_shift": None,
                        "d_close_reference": None, "d1_open_reference": None,
                        "interval_structure": None,
                        "execution": {
                            "status": "data_unavailable",
                            "reason": "fixed_path_bar_missing_no_forward_shift",
                            "entry_date": "2026-07-30", "entry_price": None,
                            "exit_date": "2026-08-01", "exit_price": None,
                            "gross_return": None, "net_return": None,
                            "market_benchmark_net_return": None, "net_excess_return": None,
                            "cost_drag": None, "daily_bar_model_limited": None,
                            "cost_model_version": "ashare-cost-v1", "cost_profile_id": "base",
                        },
                    },
                ],
            }],
        },
    }


def _run() -> MarketScanRun:
    return MarketScanRun(
        id=12,
        status="success",
        trigger="manual",
        mode="official",
        rule_version="full-market-scan-v3:abcdef12",
        as_of="2026-07-29 16:00:00",
        data_date="2026-07-29",
        quote_date="2026-07-29",
        scope=FULL_MARKET_SCOPE,
        stock_pool_source="akshare",
        total_count=1,
        excluded_count=0,
        processed_count=1,
        success_count=1,
        missing_count=0,
        skipped_count=0,
        retry_count=0,
        progress_pct=100,
        coverage_pct=100,
        created_at="2026-07-29 16:00:00",
        updated_at="2026-07-29 16:10:00",
        started_at="2026-07-29 16:00:01",
        finished_at="2026-07-29 16:10:00",
        duration_ms=599_000,
        snapshot_digest="a" * 64,
        snapshot_seal_origin="publication",
        snapshot_sealed_at="2026-07-29 16:10:00",
    )


def _item() -> MarketScanResultItem:
    return MarketScanResultItem(
        run_id=12,
        symbol="000001.SZ",
        code="000001",
        market="SZ",
        name='=HYPERLINK("https://example.invalid")',
        industry="+银行",
        list_date="1991-04-03",
        status="success",
        rank=1,
        score=92,
        raw_score=91.123456,
        trend_score=88,
        leader_score=77,
        data_quality_score=96,
        price=12.34,
        change_pct=1.25,
        turnover_rate=2.5,
        volume_ratio=1.1,
        amount=1_234_567_890.12,
        tags=["放量", "突破"],
        reason="@趋势和量价配合",
        error=None,
        data_date="2026-07-29",
        quote_timestamp="2026-07-29 15:00:00",
        quote_observed_at="2026-07-29T07:00:00Z",
        quote_source="tencent",
        kline_source="akshare",
        adjustment_mode="qfq",
        degradation_reasons=["-provider_reason"],
        updated_at="2026-07-29 16:08:00",
        score_details={
            "run_rule_version": "full-market-scan-v3:abcdef12",
            "score_spec_hash": "abcdef123456",
            "components": {
                "leader_score": {"base": 50, "trend_delta": 4, "rule_deltas": {"amount": 2.5}},
                "final_score": {"quality_penalty": 0.6, "base": 91.4, "rank_discount": 0.276544},
                "rank_refinement": {"score": 0.445, "weighted_terms": {"ma_alignment": 0.2}},
            },
            "ranking": {
                "tie_break": [["raw_score", "desc"], ["symbol", "asc"]],
                "tie_break_values": {"raw_score": 91.123456, "symbol": "000001.SZ"},
            },
        },
    )
